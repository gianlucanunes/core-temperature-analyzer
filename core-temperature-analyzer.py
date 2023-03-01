# ----------------------------------------------------------

# Core Temperature Analyzer
# by Gianluca Nunes
# March 1, 2023
#
# This program analyzes the core temperature of a computer
# using the software Core Temp. Core Temp creates the log files 
# (.csv) and this program organizes the data from the log files 
# in order to make a more intuitive chart with the Date and its
# measures. After that, it returns you with an average CPU
# temperature of the last five days.
#
# ----------------------------------------------------------


# Importing libraries
import datetime
import numpy as np
import pandas as pd
import time
import os
import openpyxl
from termcolor import colored


# Header
os.system('cls' if os.name == 'nt' else 'clear')

print("=" * 50)
print("{:^50}".format("Core Temperature Analyzer"))
print("{:^50}".format("by Gianluca Nunes"))
print("=" * 50)


# Warning
print(colored("\n\nWarning: before using the program, make sure Core Temp is closed and not running in your background.\n", "red", "on_white", attrs=["bold"]))
input("\nWhen ready, press any key to start the Core Temperature Analyzer.\nRemember: you should not open any file related to this program during the analysis process.\n")

time.sleep(1)
print("Please wait...")
time.sleep(1)


# Organizes the measured temperatures inside an excel file called temp.xlsx
def write_spreadsheet(date, temperature):
    # Load the file or creates one if it does not exist
    spreadsheet = openpyxl.load_workbook(r"TEMP_FILE_PATH") # EXAMPLE: spreadsheet = openpyxl.load_workbook(r"C:\Users\John\Documents\TempAnalyzer\logs\temp.xlsx")
    sheet = spreadsheet.active

    # Find the first column which contains the date passed as an argument
    control = 1
    found = False
    while sheet.cell(row=1, column=control).value is not None:
        if sheet.cell(row=1, column=control).value == date:
            found = True
            break
        control += 1

    # If the date is already at the column, adds the new temperature measured in the first empty cell
    if found:
        line = 2
        while sheet.cell(row=line, column=control).value is not None:
            line += 1
        sheet.cell(row=line, column=control).value = temperature

    # If the date is not in the spreadsheet, adds it to the first empty column
    else:
        control = 1
        while sheet.cell(row=1, column=control).value is not None:
            control += 1

        line = 2
        sheet.cell(row=1, column=control).value = date
        sheet.cell(row=line, column=control).value = temperature

    # Save the changes inside the spreadsheet
    spreadsheet.save(r"TEMP_FILE_PATH") # EXAMPLE: spreadsheet.save(r"C:\Users\John\Documents\TempAnalyzer\logs\temp.xlsx")



# Take the measured temperatures from the .csv log files
def take_info():
    directory = r"CORE_TEMP_FOLDER" # EXAMPLE: directory = r"C:\Users\John\Documents\Core Temp"

    # Open the Core Temp log file, which starts with "CT" (not measured yet) and ends with ".csv"
    for filename in os.listdir(directory):
        if filename.startswith("CT") and filename.endswith(".csv"):
            file_path = os.path.join(directory, filename)
            with open(file_path, "r") as file:

                # Read the lines starting from the 10th line
                lines = file.readlines()
                for i in range(10, len(lines)):
                    line = lines[i]
                    if line.startswith("Session end") or line.startswith(' '):
                        break

                    # Get the date and the temperature
                    date = line[10:17]
                    temperature = line[18:20]

                    # Write the info in the temp.xlsx file
                    write_spreadsheet(date, temperature)

            # Change the log file name to avoid duplicated measures
            new_filename = "temperatures_processed" + filename
            file_path = os.path.join(directory, filename)
            new_file_path = os.path.join(directory, new_filename)
            os.rename(file_path, new_file_path)

            # Warn the user about the progress
            print(f"The log {filename} has been processed.")

# Call the function to get the info from the log files
take_info()


# Gives the analysis using the american date standard with Fahrenheit as a measure scale
# Read the file with pandas
df = pd.read_excel(r"TEMP_FILE_PATH", header=0) # EXAMPLE: df = pd.read_excel(r"C:\Users\John\Documents\TempAnalyzer\logs\temp.xlsx", header=0)

# df.iloc[:-1] = df.iloc[:-1].astype(str)
# Get the last 5 filled columns
last_5_columns = df.columns[-5:]

# Creates a list to store the 5 measured temperatures in order to calculate an average of them
mean_temps = []

# Iterates through the column
for column in last_5_columns:

    # Get the header, which is the date
    date = str(df[column].name)

    # Get the measured temperatures in the lines below the header
    temperatures = df[column][0:]

    # Calculates the average temperature
    mean_temp = round(np.mean(temperatures), 2)

    # Convert it to Fahrenheit
    mean_temp_fahrenheit = (mean_temp * 9/5) + 32
    mean_temps.append(mean_temp_fahrenheit)

    # Print the results of each day
    print(f"Day {date}: average temperature of {round(np.mean(mean_temp_fahrenheit), 2)} ºF")

# Print the final result
final_mean_temps = round(np.mean(mean_temps), 2)
print(f"\nFinal Analysis: the average core temperature from the last 5 days is {final_mean_temps} ºF")


"""
# Gives the analysis using the brazilian date standard with Celsius as a measure scale
# Lê o arquivo com o Pandas
df = pd.read_excel(r"TEMP_FILE_PATH", header=0) # EXEMPLO: df = pd.read_excel(r"C:\Users\John\Documents\TempAnalyzer\logs\temp.xlsx", header=0)

# df.iloc[:-1] = df.iloc[:-1].astype(str)
# Pega as últimas 5 colunas preenchidas
last_5_columns = df.columns[-5:]

mean_temps = []
# Itera sobre as últimas 5 colunas
for column in last_5_columns:
    # Pega o nome da coluna, que deve ser a data
    date_string = str(df[column].name)
    # Converte a data para o formato '%m/%d/%y'
    date = datetime.datetime.strptime(date_string, '%m/%d/%y')
    # Converte a data para o padrão brasileiro (dd/mm/aaaa)
    date = date.strftime('%d/%m/%Y')

    # Pega as temperaturas nas outras linhas da coluna
    temperatures = df[column][0:]

    # Calcula a média das temperaturas
    mean_temp = round(np.mean(temperatures), 2)
    mean_temps.append(mean_temp)

    # Printa a data e a média das temperaturas
    print(f"Dia {date}: temperatura média de {mean_temp} ºC")

final_mean_temp = round(np.mean(mean_temps), 2)
print(f"\nFinal Analysis: the average core temperature from the last 5 days is {final_mean_temp} ºC")
time.sleep(1)
"""

